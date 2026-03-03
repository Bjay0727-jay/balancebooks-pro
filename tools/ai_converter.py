#!/usr/bin/env python3
"""
BalanceBooks Pro - AI Spreadsheet Converter

A Python script that converts any billing spreadsheet into BalanceBooks format
using OpenAI for intelligent column detection and transaction categorization.

Usage:
  CLI mode:   python ai_converter.py convert invoice.xlsx -o output.csv
  Server mode: python ai_converter.py serve --port 5555

Requires: pip install openai flask flask-cors openpyxl
"""

import argparse
import csv
import io
import json
import os
import re
import sys
from datetime import datetime, timedelta
from pathlib import Path

# ── BalanceBooks categories ─────────────────────────────────────────────────

CATEGORIES = [
    "income", "housing", "utilities", "groceries", "transportation",
    "healthcare", "insurance", "entertainment", "dining", "shopping",
    "subscriptions", "education", "tithes", "savings", "investment",
    "debt", "childcare", "pets", "personal", "gifts", "transfer", "other",
]

COLUMN_ROLES = [
    "date", "description", "amount", "debit", "credit",
    "category", "type", "status", "reference", "balance", "skip",
]

CATEGORY_ALIASES = {
    "tithe": "tithes", "offering": "tithes", "church": "tithes",
    "donation": "gifts", "charity": "gifts", "gift": "gifts",
    "baby": "childcare", "daycare": "childcare", "child": "childcare", "kids": "childcare",
    "pet": "pets", "dog": "pets", "cat": "pets", "vet": "pets",
    "hair": "personal", "salon": "personal", "spa": "personal",
    "gym": "healthcare", "medical": "healthcare", "doctor": "healthcare", "pharmacy": "healthcare",
    "car": "transportation", "auto": "transportation", "gas": "transportation",
    "fuel": "transportation", "uber": "transportation", "lyft": "transportation",
    "food": "groceries", "grocery": "groceries", "supermarket": "groceries", "market": "groceries",
    "restaurant": "dining", "coffee": "dining", "cafe": "dining", "takeout": "dining",
    "doordash": "dining", "grubhub": "dining", "ubereats": "dining",
    "netflix": "subscriptions", "spotify": "subscriptions", "hulu": "subscriptions",
    "subscription": "subscriptions", "disney": "subscriptions", "apple music": "subscriptions",
    "amazon": "shopping", "walmart": "shopping", "target": "shopping", "costco": "shopping",
    "rent": "housing", "mortgage": "housing", "hoa": "housing",
    "electric": "utilities", "water": "utilities", "internet": "utilities",
    "phone": "utilities", "utility": "utilities", "comcast": "utilities", "att": "utilities",
    "salary": "income", "paycheck": "income", "wages": "income", "freelance": "income",
    "deposit": "income", "direct dep": "income", "payroll": "income", "refund": "income",
    "insurance": "insurance", "premium": "insurance", "geico": "insurance", "state farm": "insurance",
    "tuition": "education", "school": "education", "textbook": "education", "university": "education",
    "invest": "investment", "stock": "investment", "dividend": "investment", "etrade": "investment",
    "loan": "debt", "credit card": "debt", "payment": "debt", "student loan": "debt",
    "saving": "savings", "transfer": "transfer", "zelle": "transfer", "venmo": "transfer",
}


# ── Date parsing ────────────────────────────────────────────────────────────

def parse_date(value):
    """Parse various date formats into YYYY-MM-DD string."""
    if value is None:
        return None
    s = str(value).strip()
    if not s:
        return None

    # Excel serial date
    try:
        num = float(s)
        if 0 < num < 100000:
            epoch = datetime(1899, 12, 30)
            dt = epoch + timedelta(days=num)
            return dt.strftime("%Y-%m-%d")
    except (ValueError, TypeError):
        pass

    # ISO: YYYY-MM-DD
    m = re.match(r"^(\d{4})-(\d{1,2})-(\d{1,2})", s)
    if m:
        return f"{m.group(1)}-{int(m.group(2)):02d}-{int(m.group(3)):02d}"

    # US: M/D/YYYY
    m = re.match(r"^(\d{1,2})/(\d{1,2})/(\d{4})$", s)
    if m:
        return f"{m.group(3)}-{int(m.group(1)):02d}-{int(m.group(2)):02d}"

    # US short: M/D/YY
    m = re.match(r"^(\d{1,2})/(\d{1,2})/(\d{2})$", s)
    if m:
        yy = int(m.group(3))
        year = (1900 + yy) if yy > 50 else (2000 + yy)
        return f"{year}-{int(m.group(1)):02d}-{int(m.group(2)):02d}"

    # US dash: M-D-YYYY
    m = re.match(r"^(\d{1,2})-(\d{1,2})-(\d{4})$", s)
    if m:
        return f"{m.group(3)}-{int(m.group(1)):02d}-{int(m.group(2)):02d}"

    # Long date: January 15, 2025
    months = ["jan", "feb", "mar", "apr", "may", "jun", "jul", "aug", "sep", "oct", "nov", "dec"]
    m = re.match(r"^([A-Za-z]+)\s+(\d{1,2}),?\s+(\d{4})$", s)
    if m:
        month_str = m.group(1).lower()[:3]
        if month_str in months:
            mi = months.index(month_str) + 1
            return f"{m.group(3)}-{mi:02d}-{int(m.group(2)):02d}"

    # Fallback: try Python's dateutil-style parsing
    for fmt in ["%m/%d/%Y", "%m-%d-%Y", "%Y/%m/%d", "%d/%m/%Y", "%B %d, %Y", "%b %d, %Y"]:
        try:
            dt = datetime.strptime(s, fmt)
            return dt.strftime("%Y-%m-%d")
        except ValueError:
            pass

    return None


def parse_amount(value):
    """Parse amount string handling $, commas, parentheses for negatives."""
    if value is None:
        return None
    s = str(value).strip()
    if not s:
        return None

    # Parentheses mean negative: (123.45) -> -123.45
    paren = re.match(r"^\(([^)]+)\)$", s)
    if paren:
        cleaned = re.sub(r"[$,\s]", "", paren.group(1))
        try:
            return -abs(float(cleaned))
        except ValueError:
            return None

    cleaned = re.sub(r"[$,\s]", "", s)
    try:
        return float(cleaned)
    except ValueError:
        return None


# ── Heuristic fallback logic ────────────────────────────────────────────────

def heuristic_detect_columns(headers):
    """Detect column roles from header names (fallback when no AI)."""
    roles = ["skip"] * len(headers)
    lower = [str(h or "").lower().strip() for h in headers]

    # Date
    for i, h in enumerate(lower):
        if h in ("date", "transaction date", "trans date", "posting date", "post date"):
            roles[i] = "date"
            break
    else:
        for i, h in enumerate(lower):
            if "date" in h:
                roles[i] = "date"
                break

    # Description
    desc_keywords = ("description", "desc", "memo", "payee", "merchant", "name",
                     "details", "narrative", "particulars")
    for i, h in enumerate(lower):
        if roles[i] != "skip":
            continue
        if h in desc_keywords or any(k in h for k in ("desc", "memo", "payee", "merchant", "narrat")):
            roles[i] = "description"
            break

    # Amount
    amt_idx = -1
    for i, h in enumerate(lower):
        if roles[i] != "skip":
            continue
        if h in ("amount", "sum", "total", "transaction amount"):
            roles[i] = "amount"
            amt_idx = i
            break

    # Debit / Credit
    if amt_idx < 0:
        debit_idx = credit_idx = -1
        for i, h in enumerate(lower):
            if roles[i] != "skip":
                continue
            if h in ("debit", "debits", "withdrawal", "withdrawals", "charge", "charges", "money out"):
                debit_idx = i
            elif h in ("credit", "credits", "deposit", "deposits", "money in"):
                credit_idx = i
        if debit_idx >= 0:
            roles[debit_idx] = "debit"
        if credit_idx >= 0:
            roles[credit_idx] = "credit"

    # Category
    for i, h in enumerate(lower):
        if roles[i] != "skip":
            continue
        if h in ("category", "cat", "class", "tag", "group"):
            roles[i] = "category"
            break

    # Type
    for i, h in enumerate(lower):
        if roles[i] != "skip":
            continue
        if h in ("type", "transaction type", "trans type"):
            roles[i] = "type"
            break

    # Status
    for i, h in enumerate(lower):
        if roles[i] != "skip":
            continue
        if h in ("paid", "status", "cleared", "reconciled"):
            roles[i] = "status"
            break

    # Balance
    for i, h in enumerate(lower):
        if roles[i] != "skip":
            continue
        if h in ("balance", "running balance", "running total"):
            roles[i] = "balance"
            break

    return roles


def heuristic_map_category(text):
    """Map description/category text to a BalanceBooks category using aliases."""
    if not text:
        return "other"
    lower = text.lower().strip()

    # Exact match
    if lower in CATEGORIES:
        return lower

    # Alias match
    for alias, cat_id in CATEGORY_ALIASES.items():
        if alias in lower:
            return cat_id

    return "other"


# ── File parsing ────────────────────────────────────────────────────────────

def read_csv_file(filepath):
    """Read CSV file and return headers + rows."""
    with open(filepath, "r", encoding="utf-8-sig") as f:
        # Detect delimiter
        sample = f.read(4096)
        f.seek(0)
        first_line = sample.split("\n")[0]
        delimiter = "\t" if "\t" in first_line and "," not in first_line else ","

        reader = csv.reader(f, delimiter=delimiter)
        rows = list(reader)

    rows = [r for r in rows if any(cell.strip() for cell in r)]
    if len(rows) < 2:
        return None, None
    return rows[0], rows[1:]


def read_excel_file(filepath):
    """Read XLS/XLSX file and return headers + rows."""
    try:
        import openpyxl
    except ImportError:
        print("Error: openpyxl required for Excel files. Install: pip install openpyxl", file=sys.stderr)
        return None, None

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append([str(cell) if cell is not None else "" for cell in row])
    wb.close()

    rows = [r for r in rows if any(cell.strip() for cell in r)]
    if len(rows) < 2:
        return None, None
    return rows[0], rows[1:]


def read_file(filepath):
    """Read CSV, TSV, XLS, or XLSX file."""
    ext = Path(filepath).suffix.lower()
    if ext in (".csv", ".tsv", ".txt"):
        return read_csv_file(filepath)
    elif ext in (".xlsx", ".xls"):
        return read_excel_file(filepath)
    else:
        print(f"Unsupported file type: {ext}", file=sys.stderr)
        return None, None


def read_file_bytes(file_bytes, filename):
    """Read from file bytes (for server uploads)."""
    ext = Path(filename).suffix.lower()
    if ext in (".csv", ".tsv", ".txt"):
        text = file_bytes.decode("utf-8-sig")
        first_line = text.split("\n")[0]
        delimiter = "\t" if "\t" in first_line and "," not in first_line else ","
        reader = csv.reader(io.StringIO(text), delimiter=delimiter)
        rows = [r for r in reader if any(cell.strip() for cell in r)]
        if len(rows) < 2:
            return None, None
        return rows[0], rows[1:]
    elif ext in (".xlsx", ".xls"):
        try:
            import openpyxl
        except ImportError:
            return None, None
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        ws = wb.active
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append([str(cell) if cell is not None else "" for cell in row])
        wb.close()
        rows = [r for r in rows if any(cell.strip() for cell in r)]
        if len(rows) < 2:
            return None, None
        return rows[0], rows[1:]
    return None, None


# ── OpenAI AI functions ─────────────────────────────────────────────────────

def get_openai_client(api_key):
    """Create OpenAI client."""
    try:
        from openai import OpenAI
        return OpenAI(api_key=api_key)
    except ImportError:
        print("Error: openai package required. Install: pip install openai", file=sys.stderr)
        return None


def ai_detect_columns(headers, sample_rows, api_key, model="gpt-4o-mini"):
    """Use AI to detect column roles and sign convention."""
    client = get_openai_client(api_key)
    if not client:
        return None

    sample_text = ""
    for i, row in enumerate(sample_rows[:3]):
        sample_text += f"  Row {i+1}: {json.dumps(row)}\n"

    prompt = f"""Analyze this financial spreadsheet and determine what each column represents.

Headers: {json.dumps(headers)}
Sample data:
{sample_text}
Allowed column roles: {json.dumps(COLUMN_ROLES)}

Rules:
- Exactly one column should be "date" (the transaction date)
- Exactly one column should be "description" (merchant name, memo, payee, etc.)
- Use "amount" for a single combined amount column, OR use "debit"/"credit" for split columns
- "balance" columns contain running account balance (should be skipped in import)
- Mark unneeded columns as "skip"
- Also determine the sign convention: "auto" (positive=income, negative=expense), "bank" (same as auto), "credit-card" (positive=charge, negative=payment), or "all-positive"

Return ONLY valid JSON in this exact format:
{{"columns": [{{"index": 0, "role": "date", "confidence": 0.95}}, ...], "sign_convention": "bank", "sign_confidence": 0.9}}"""

    try:
        response = client.chat.completions.create(
            model=model,
            messages=[
                {"role": "system", "content": "You are a financial data analyst. Always respond with valid JSON only, no markdown fences."},
                {"role": "user", "content": prompt},
            ],
            temperature=0.1,
            max_tokens=1000,
        )
        content = response.choices[0].message.content.strip()
        # Strip markdown code fences if present
        content = re.sub(r"^```(?:json)?\s*", "", content)
        content = re.sub(r"\s*```$", "", content)
        result = json.loads(content)

        # Validate roles
        columns = result.get("columns", [])
        for col in columns:
            if col.get("role") not in COLUMN_ROLES:
                col["role"] = "skip"
                col["confidence"] = 0.0

        return result
    except Exception as e:
        print(f"AI column detection failed: {e}", file=sys.stderr)
        return None


def ai_categorize(descriptions, api_key, model="gpt-4o-mini", batch_size=30):
    """Use AI to categorize transaction descriptions in batches."""
    client = get_openai_client(api_key)
    if not client:
        return None

    all_results = []

    for batch_start in range(0, len(descriptions), batch_size):
        batch = descriptions[batch_start:batch_start + batch_size]
        numbered = "\n".join(f"  {i}: {desc}" for i, desc in enumerate(batch))

        prompt = f"""Classify each transaction description into exactly one BalanceBooks category.

Categories: {", ".join(CATEGORIES)}

Rules:
- "income" is for salary, wages, deposits, refunds, interest earned
- "tithes" is for church, tithe, offering donations
- "transfer" is for transfers between accounts (Zelle, Venmo, bank transfer)
- "other" is the fallback when uncertain
- Return confidence 0.0-1.0 for each classification

Descriptions:
{numbered}

Return ONLY a JSON array: [{{"i": 0, "c": "groceries", "p": 0.95}}, ...]
where i=index, c=category id, p=confidence probability"""

        try:
            response = client.chat.completions.create(
                model=model,
                messages=[
                    {"role": "system", "content": "You are a financial transaction classifier. Always respond with valid JSON only, no markdown fences."},
                    {"role": "user", "content": prompt},
                ],
                temperature=0.1,
                max_tokens=2000,
            )
            content = response.choices[0].message.content.strip()
            content = re.sub(r"^```(?:json)?\s*", "", content)
            content = re.sub(r"\s*```$", "", content)
            batch_results = json.loads(content)

            for item in batch_results:
                idx = item.get("i", 0)
                cat = item.get("c", "other")
                conf = item.get("p", 0.5)
                if cat not in CATEGORIES:
                    cat = "other"
                    conf = 0.0
                all_results.append({
                    "index": batch_start + idx,
                    "category": cat,
                    "confidence": min(1.0, max(0.0, conf)),
                })
        except Exception as e:
            print(f"AI categorization batch failed: {e}", file=sys.stderr)
            # Fallback for this batch: use heuristic
            for i, desc in enumerate(batch):
                all_results.append({
                    "index": batch_start + i,
                    "category": heuristic_map_category(desc),
                    "confidence": 0.3,
                })

    return all_results


# ── Conversion logic ────────────────────────────────────────────────────────

def convert_spreadsheet(headers, data_rows, api_key=None, model="gpt-4o-mini",
                        sign_convention="auto", default_paid=True):
    """Convert parsed spreadsheet data into BalanceBooks transactions."""
    # Step 1: Detect columns
    if api_key:
        ai_result = ai_detect_columns(headers, data_rows[:3], api_key, model)
        if ai_result:
            roles = ["skip"] * len(headers)
            for col in ai_result.get("columns", []):
                idx = col.get("index", -1)
                if 0 <= idx < len(roles):
                    roles[idx] = col.get("role", "skip")
            suggested_sign = ai_result.get("sign_convention", sign_convention)
            if suggested_sign in ("auto", "bank", "credit-card", "all-positive"):
                sign_convention = suggested_sign
        else:
            roles = heuristic_detect_columns(headers)
    else:
        roles = heuristic_detect_columns(headers)

    # Find column indices
    date_idx = roles.index("date") if "date" in roles else -1
    desc_idx = roles.index("description") if "description" in roles else -1
    amt_idx = roles.index("amount") if "amount" in roles else -1
    debit_idx = roles.index("debit") if "debit" in roles else -1
    credit_idx = roles.index("credit") if "credit" in roles else -1
    cat_idx = roles.index("category") if "category" in roles else -1
    type_idx = roles.index("type") if "type" in roles else -1
    status_idx = roles.index("status") if "status" in roles else -1

    if date_idx < 0 or desc_idx < 0:
        return [], ["Could not detect Date and Description columns"], roles

    if amt_idx < 0 and debit_idx < 0 and credit_idx < 0:
        return [], ["Could not detect any amount column"], roles

    # Step 2: Parse transactions
    transactions = []
    errors = []
    descriptions_for_ai = []
    desc_indices = []  # maps AI desc index to transaction index

    for row_num, row in enumerate(data_rows, start=2):
        if not row or all(not str(c).strip() for c in row):
            continue

        # Date
        date_raw = row[date_idx] if date_idx < len(row) else ""
        date = parse_date(date_raw)
        if not date:
            errors.append(f"Row {row_num}: Invalid date \"{date_raw}\"")
            continue

        # Description
        desc = str(row[desc_idx] if desc_idx < len(row) else "").strip()
        if not desc:
            errors.append(f"Row {row_num}: Missing description")
            continue

        # Amount
        amount = 0.0
        if amt_idx >= 0:
            amount = parse_amount(row[amt_idx] if amt_idx < len(row) else "")
            if amount is None:
                errors.append(f"Row {row_num}: Invalid amount")
                continue
        else:
            debit_val = parse_amount(row[debit_idx] if debit_idx >= 0 and debit_idx < len(row) else "")
            credit_val = parse_amount(row[credit_idx] if credit_idx >= 0 and credit_idx < len(row) else "")
            if debit_val and debit_val != 0:
                amount = -abs(debit_val)
            elif credit_val and credit_val != 0:
                amount = abs(credit_val)
            else:
                errors.append(f"Row {row_num}: No valid amount")
                continue

        if amount == 0:
            errors.append(f"Row {row_num}: Zero amount")
            continue

        # Type column
        type_val = str(row[type_idx] if type_idx >= 0 and type_idx < len(row) else "").lower().strip()
        cat_val = str(row[cat_idx] if cat_idx >= 0 and cat_idx < len(row) else "").strip()

        # Determine income vs expense
        is_income = False
        if type_val in ("income", "credit", "deposit"):
            is_income = True
        elif type_val in ("expense", "debit", "charge"):
            is_income = False
        elif sign_convention == "credit-card":
            is_income = amount < 0
            amount = abs(amount)
            if not is_income:
                amount = -amount
        elif sign_convention == "all-positive":
            mapped = heuristic_map_category(cat_val or desc)
            is_income = mapped == "income"
        else:
            is_income = amount > 0

        final_amount = abs(amount) if is_income else -abs(amount)

        # Category (heuristic for now, AI override below)
        if is_income:
            category = "income"
        elif cat_val:
            category = heuristic_map_category(cat_val)
        else:
            category = heuristic_map_category(desc)

        # Paid status
        paid = default_paid
        if status_idx >= 0 and status_idx < len(row):
            paid_val = str(row[status_idx]).lower().strip()
            if paid_val in ("yes", "1", "true", "y", "paid", "cleared", "reconciled"):
                paid = True
            elif paid_val in ("no", "0", "false", "n", "unpaid", "pending"):
                paid = False

        tx = {
            "date": date,
            "desc": desc,
            "amount": round(final_amount, 2),
            "category": category,
            "paid": paid,
        }
        transactions.append(tx)

        # Collect descriptions for AI categorization (only for non-income, no cat column)
        if not is_income and cat_idx < 0:
            descriptions_for_ai.append(desc)
            desc_indices.append(len(transactions) - 1)

    # Step 3: AI categorization
    if api_key and descriptions_for_ai:
        ai_cats = ai_categorize(descriptions_for_ai, api_key, model)
        if ai_cats:
            for result in ai_cats:
                ai_idx = result["index"]
                if ai_idx < len(desc_indices):
                    tx_idx = desc_indices[ai_idx]
                    if result["confidence"] > 0.5:
                        transactions[tx_idx]["category"] = result["category"]
                        transactions[tx_idx]["ai_confidence"] = result["confidence"]

    # Sort by date descending
    transactions.sort(key=lambda t: t["date"], reverse=True)

    return transactions, errors, roles


# ── Output formatters ───────────────────────────────────────────────────────

def to_csv_string(transactions):
    """Convert transactions to BalanceBooks CSV string."""
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Date", "Description", "Amount", "Category", "Type", "Paid"])
    for tx in transactions:
        writer.writerow([
            tx["date"],
            tx["desc"],
            f"{tx['amount']:.2f}",
            tx["category"],
            "income" if tx["amount"] >= 0 else "expense",
            "yes" if tx["paid"] else "no",
        ])
    return output.getvalue()


def to_json(transactions):
    """Convert transactions to JSON-serializable list."""
    return [{
        "date": tx["date"],
        "desc": tx["desc"],
        "amount": tx["amount"],
        "category": tx["category"],
        "paid": tx["paid"],
        "ai_confidence": tx.get("ai_confidence"),
    } for tx in transactions]


# ── CLI mode ────────────────────────────────────────────────────────────────

def cli_convert(args):
    """CLI: Convert a spreadsheet file."""
    headers, rows = read_file(args.input)
    if headers is None:
        print("Error: Could not read file or file is empty.", file=sys.stderr)
        sys.exit(1)

    api_key = args.api_key or os.environ.get("OPENAI_API_KEY")
    model = args.model or "gpt-4o-mini"

    print(f"Read {len(rows)} rows with {len(headers)} columns")
    if api_key:
        print(f"Using AI ({model}) for smart detection and categorization")
    else:
        print("No API key - using heuristic pattern matching")

    transactions, errors, roles = convert_spreadsheet(
        headers, rows,
        api_key=api_key,
        model=model,
        sign_convention=args.sign or "auto",
        default_paid=not args.unpaid,
    )

    print(f"Converted {len(transactions)} transactions ({len(errors)} errors)")

    if errors and args.verbose:
        for err in errors[:10]:
            print(f"  Warning: {err}", file=sys.stderr)
        if len(errors) > 10:
            print(f"  ... and {len(errors) - 10} more", file=sys.stderr)

    # Write output
    output_path = args.output
    if not output_path:
        stem = Path(args.input).stem
        output_path = f"{stem}-balancebooks.csv"

    csv_content = to_csv_string(transactions)
    with open(output_path, "w", encoding="utf-8", newline="") as f:
        f.write(csv_content)

    print(f"Saved to: {output_path}")

    # Print summary
    income = sum(tx["amount"] for tx in transactions if tx["amount"] > 0)
    expenses = sum(abs(tx["amount"]) for tx in transactions if tx["amount"] < 0)
    print(f"  Income:   ${income:,.2f}")
    print(f"  Expenses: ${expenses:,.2f}")
    print(f"  Net:      ${income - expenses:,.2f}")


# ── Flask server mode ───────────────────────────────────────────────────────

def create_server():
    """Create Flask app for server mode."""
    try:
        from flask import Flask, request, jsonify
        from flask_cors import CORS
    except ImportError:
        print("Error: Flask required for server mode. Install: pip install flask flask-cors",
              file=sys.stderr)
        sys.exit(1)

    app = Flask(__name__)
    CORS(app)

    @app.route("/health", methods=["GET"])
    def health():
        return jsonify({"status": "ok", "version": "1.0.0"})

    @app.route("/analyze", methods=["POST"])
    def analyze():
        """Analyze spreadsheet columns using AI."""
        data = request.get_json()
        if not data:
            return jsonify({"error": "No JSON body"}), 400

        headers = data.get("headers", [])
        sample_rows = data.get("sample_rows", [])
        api_key = data.get("api_key") or os.environ.get("OPENAI_API_KEY")
        model = data.get("model", "gpt-4o-mini")

        if not headers:
            return jsonify({"error": "No headers provided"}), 400

        if api_key:
            result = ai_detect_columns(headers, sample_rows, api_key, model)
            if result:
                return jsonify(result)

        # Fallback to heuristic
        roles = heuristic_detect_columns(headers)
        columns = [{"index": i, "role": r, "confidence": 0.7 if r != "skip" else 0.0}
                    for i, r in enumerate(roles)]
        return jsonify({
            "columns": columns,
            "sign_convention": "auto",
            "sign_confidence": 0.5,
            "fallback": True,
        })

    @app.route("/categorize", methods=["POST"])
    def categorize():
        """Categorize transaction descriptions using AI."""
        data = request.get_json()
        if not data:
            return jsonify({"error": "No JSON body"}), 400

        descriptions = data.get("descriptions", [])
        api_key = data.get("api_key") or os.environ.get("OPENAI_API_KEY")
        model = data.get("model", "gpt-4o-mini")

        if not descriptions:
            return jsonify({"error": "No descriptions provided"}), 400

        if api_key:
            results = ai_categorize(descriptions, api_key, model)
            if results:
                return jsonify({"results": results})

        # Fallback to heuristic
        results = [{"index": i, "category": heuristic_map_category(d), "confidence": 0.3}
                    for i, d in enumerate(descriptions)]
        return jsonify({"results": results, "fallback": True})

    @app.route("/convert", methods=["POST"])
    def convert():
        """Full conversion: upload file, return BalanceBooks transactions."""
        if "file" not in request.files:
            return jsonify({"error": "No file uploaded"}), 400

        file = request.files["file"]
        api_key = request.form.get("api_key") or os.environ.get("OPENAI_API_KEY")
        model = request.form.get("model", "gpt-4o-mini")
        sign = request.form.get("sign_convention", "auto")

        file_bytes = file.read()
        headers, rows = read_file_bytes(file_bytes, file.filename)
        if headers is None:
            return jsonify({"error": "Could not parse file"}), 400

        transactions, errors, roles = convert_spreadsheet(
            headers, rows,
            api_key=api_key,
            model=model,
            sign_convention=sign,
        )

        income = sum(tx["amount"] for tx in transactions if tx["amount"] > 0)
        expenses = sum(abs(tx["amount"]) for tx in transactions if tx["amount"] < 0)

        return jsonify({
            "transactions": to_json(transactions),
            "errors": errors[:50],
            "column_roles": roles,
            "stats": {
                "total": len(transactions),
                "income": round(income, 2),
                "expenses": round(expenses, 2),
                "net": round(income - expenses, 2),
            },
        })

    return app


# ── Main entry point ────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="BalanceBooks Pro AI Spreadsheet Converter",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Convert with AI (uses OPENAI_API_KEY env var)
  python ai_converter.py convert bank_statement.xlsx

  # Convert with explicit API key
  python ai_converter.py convert invoice.csv --api-key sk-...

  # Convert without AI (heuristic only)
  python ai_converter.py convert data.csv

  # Start HTTP server for web app integration
  python ai_converter.py serve --port 5555
        """,
    )
    subparsers = parser.add_subparsers(dest="command")

    # Convert command
    convert_parser = subparsers.add_parser("convert", help="Convert a spreadsheet file")
    convert_parser.add_argument("input", help="Input file (CSV, TSV, XLS, XLSX)")
    convert_parser.add_argument("-o", "--output", help="Output CSV file path")
    convert_parser.add_argument("--api-key", help="OpenAI API key (or set OPENAI_API_KEY env var)")
    convert_parser.add_argument("--model", default="gpt-4o-mini", help="AI model (default: gpt-4o-mini)")
    convert_parser.add_argument("--sign", choices=["auto", "bank", "credit-card", "all-positive"],
                                default="auto", help="Amount sign convention")
    convert_parser.add_argument("--unpaid", action="store_true", help="Mark all transactions as unpaid")
    convert_parser.add_argument("-v", "--verbose", action="store_true", help="Show detailed errors")

    # Serve command
    serve_parser = subparsers.add_parser("serve", help="Start HTTP server for web app integration")
    serve_parser.add_argument("--port", type=int, default=5555, help="Server port (default: 5555)")
    serve_parser.add_argument("--host", default="127.0.0.1", help="Server host (default: 127.0.0.1)")

    args = parser.parse_args()

    if args.command == "convert":
        cli_convert(args)
    elif args.command == "serve":
        app = create_server()
        print(f"BalanceBooks AI Converter server starting on http://{args.host}:{args.port}")
        app.run(host=args.host, port=args.port, debug=False)
    else:
        parser.print_help()


if __name__ == "__main__":
    main()
